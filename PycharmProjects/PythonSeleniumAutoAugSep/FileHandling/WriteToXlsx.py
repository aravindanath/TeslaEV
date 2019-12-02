
import openpyxl


o = openpyxl.load_workbook("./TestData.xlsx")

o.create_sheet("Demo",0)
vk  = o.sheetnames
print(vk)
#
# sh1=o["URLs"]
# sh1['A2']="www.amazon.in"

sh = o["Demo"]
sh['A2']="https://amazon.com"

o.save("./TestData.xlsx")
